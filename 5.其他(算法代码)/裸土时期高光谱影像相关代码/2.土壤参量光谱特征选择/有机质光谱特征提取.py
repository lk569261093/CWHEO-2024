import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from scipy.optimize import differential_evolution
from openpyxl import Workbook
from tqdm import tqdm
import time

def read_excel_data(file_name):
    df = pd.read_excel(file_name)
    df = df.dropna()
    return df

def feature_selection(df, target_col, n_features=25):
    X = df.iloc[:, 3:].values  # 修改为从第4列开始取特征
    y = df[target_col].values

    rf = RandomForestRegressor(n_estimators=100, random_state=42)

    def objective_function(weights):
        selected_features = X[:, weights > 0.5]
        rf.fit(selected_features, y)
        feature_importances = rf.feature_importances_
        return -np.sum(feature_importances)

    bounds = [(0, 1)] * len(df.columns[3:])  # 修改为从第4列开始计算特征数量

    result = differential_evolution(objective_function, bounds, maxiter=100, seed=42)
    best_weights = (result.x > 0.5)

    rf.fit(X[:, best_weights], y)
    feature_importances = rf.feature_importances_
    top_n_features = np.argsort(feature_importances)[-n_features:]
    return df.columns[3:][best_weights][top_n_features]  # 修改为从第4列开始取特征名称

def save_selected_features_to_excel(file_name, df, target_col, selected_features):
    wb = Workbook()
    ws = wb.active

    # 写入目标列
    ws.cell(row=1, column=1, value=target_col)
    for i, value in enumerate(df[target_col].values):
        ws.cell(row=i+2, column=1, value=value)

    # 写入选择的特征
    for j, feature in enumerate(selected_features):
        ws.cell(row=1, column=j+2, value=feature)
        for i, value in enumerate(df[feature].values):
            ws.cell(row=i+2, column=j+2, value=value)

    wb.save(file_name)

def main(input_file_name):
    start_time = time.time()
    df = read_excel_data(input_file_name)
    nutrients = ['有机质(g/kg)', '含盐量(g/kg)']
    output_files = ['2.有机质光谱特征.xlsx', '2.含盐量光谱特征.xlsx']

    for nutrient, output_file in zip(tqdm(nutrients, desc='Feature Selection Progress'), output_files):
        selected_features = feature_selection(df, nutrient)
        save_selected_features_to_excel(output_file, df, nutrient, selected_features)

    end_time = time.time()
    elapsed_time = end_time - start_time
    hours, rem = divmod(elapsed_time, 3600)
    minutes, seconds = divmod(rem, 60)
    print(f"程序运行时间：{int(hours):02}:{int(minutes):02}:{seconds:.2f}")

if __name__ == '__main__':
    input_file_name = '化验数据及对应反射率数据.xlsx'
    main(input_file_name)
